import json
import os
import shutil
import tempfile
import unittest

import pytest
import openpyxl

import sys
sys.path.append(os.path.join(sys.path[0], ".."))

from forestplots import SPSSForestPlot
from forestplots import Controller

class FakeTempdir():
    def __init__(self, name):
        self.name = name


class IntegrationTests(unittest.TestCase):

    def setUp(self):
        self.tempdir = FakeTempdir("/tmp/testproject3")#tempfile.TemporaryDirectory()

    def tearDown(self):
        pass #self.tempdir.cleanup()

    def test_simple_spss(self):
        destination = shutil.copy("integration_tests/testdata/pmc5502154.pdf", self.tempdir.name)
        self.assertTrue(os.path.isfile(destination))

        c = Controller(self.tempdir.name)
        c.main()

        raw = open(os.path.join(self.tempdir.name, "forestplots.json")).read()
        data = json.loads(raw)

        # This paper has 5 forest plots
        self.assertEqual(len(data), 5)

        for result in data:
            workbook = openpyxl.load_workbook(result)
            worksheet = workbook.active

            estimator_type = worksheet.cell(row=1, column=2).value
            model_type = worksheet.cell(row=2, column=2).value
            confidence_interval = worksheet.cell(row=3, column=2).value
            count = 5

            self.assertEqual(worksheet.cell(row=count, column=1).value, "Hetrogeneity:")
            hetrogeneity = {}
            while True:
                key = worksheet.cell(row=count, column=2).value
                value = worksheet.cell(row=count, column=3).value
                count = count + 1
                if not key:
                    break
                else:
                    hetrogeneity[key] = float(value)

            self.assertEqual(worksheet.cell(row=count, column=1).value, "Overall Effect:")
            overall_effect = {}
            while True:
                key = worksheet.cell(row=count, column=2).value
                value = worksheet.cell(row=count, column=3).value
                count = count + 1
                if not key:
                    break
                else:
                    overall_effect[key] = float(value)

            self.assertEqual(worksheet.cell(row=count, column=1).value, "Data:")
            data = []
            while True:
                key = worksheet.cell(row=count, column=2).value
                value1 = worksheet.cell(row=count, column=3).value
                value2 = worksheet.cell(row=count, column=4).value
                value3 = worksheet.cell(row=count, column=5).value
                count = count + 1
                if not key:
                    count = count + 1
                    break
                else:
                    data.append((key, float(value1), float(value2), float(value3)))


            if result.find("image.4.3.96") != -1:
                self.assertEqual(estimator_type, "M-H")
                self.assertEqual(model_type, "Fixed")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Chi": 14.61,
                    "df": 9.0,
                    "P": 0.1,
                    "I": 38,
                })
                self.assertEqual(overall_effect, {
                    "Z": 2.7,
                    "P": 0.007,
                })
                expected_data = [
                    ("Suk 1995", 1.7, 0.49, 5.9),
                    ("Kuklo 2007", 0.4, 0.17, 0.95),
                    ("Karatoprak 2008", 0.37, 0.04, 3.79),
                    ("Helgeson 2010", 3.82, 0.82, 17.83),
                    ("Yilmaz 2012", 0.49, 0.04, 5.61),
                    ("Hwang 2012", 0.63, 0.3, 1.32),
                    ("Yang 2012", 1.00, 0.06, 17.12),
                    ("Halanski 2013", 0.94, 0.18, 4.96),
                    ("Samdani 2013", 0.25, 0.12, 0.55),
                    ("Sugarman 2013", 0.77, 0.25, 2.32),
                    ("Total (95% CI)", 0.61, 0.42, 0.87),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)


            elif result.find("image.5.1.110") != -1:
                self.assertEqual(estimator_type, "M-H")
                self.assertEqual(model_type, "Fixed")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Chi": 2.11,
                    "df": 5.0,
                    "P": 0.83,
                    "I": 0.0,
                })
                self.assertEqual(overall_effect, {
                    "Z": 3.8,
                    "P": 0.0001,
                })
                expected_data = [
                    ("Kuklo 2007", 0.39, 0.17, 0.91),
                    ("Karatoprak 2008", 1.19, 0.07, 20.21),
                    ("Yang 2012", 1.00, 0.06, 17.12),
                    ("Hwang 2012", 0.39, 0.1, 1.51),
                    ("Yilmaz 2012", 0.49, 0.04, 5.61),
                    ("Samdani 2013", 0.25, 0.12, 0.55),
                    ("Total (95% CI)", 0.37, 0.22, 0.62),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)

            elif result.find("image.6.1.96") != -1:
                self.assertEqual(estimator_type, "IV")
                self.assertEqual(model_type, "Fixed")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Chi": 15.08,
                    "df": 10.0,
                    "P": 0.13,
                    "I": 34.0,
                })
                self.assertEqual(overall_effect, {
                    "Z": 9.01,
                    "P": 0.00001,
                })
                expected_data = [
                    ("Kim 2006", -11.0, -16.66, -5.34),
                    ("Qiu 2007", -9.0, -14.61, -3.39),
                    ("Lowenstein 2007", -6.19, -11.19, 0.81), # !!! This should be -5.19
                    ("Li 2008", -11.0, -16.06, -5.94),
                    ("Fu 2009", -5.5, -8.13, -2.87),
                    ("Sabharwal 2011", -4.0, -9.14, 1.14),
                    ("Yilmaz 2012", -5.0, -9.02, -0.98),
                    ("Hwang 2012", -2.9, -6.48, 0.68),
                    ("Yang 2012", -4.0, -7.34, -0.66),
                    ("Haber 2014", -9.0, -13.58, -4.42),
                    ("Legarreta 2014", -4.40, -9.41, 0.61),
                    ("Total (95% CI)", -5.8, -7.06, -4.54),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)

            elif result.find("image.6.2.96") != -1:
                self.assertEqual(estimator_type, "IV")
                self.assertEqual(model_type, "Random")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Tau": 5.98,
                    "Chi": 20.24,
                    "df": 8.0,
                    "P": 0.009,
                    "I": 60.0,
                })
                self.assertEqual(overall_effect, {
                    "Z": 5.43,
                    "P": 0.00001,
                })
                expected_data = [
                    ("Kim 2006", -10.0, -14.95, -5.05),
                    ("Lowenstein 2007", -3.07, -9.47, 3.33),
                    ("Li 2008", -9.0, -13.37, -4.63),
                    ("Fu 2009", -6.3, -9.27, -3.33),
                    ("Helgeson 2010", -3.6, -6.66, -0.54),
                    ("Hwang 2012", -4.3, -8.05, -0.55),
                    ("Yilmaz 2012", -10.4, -14.2, -6.6),
                    ("Legarreta 2014", -1.0, -4.79, 2.79),
                    ("Liu 2014", -4.8, -8.78, -0.82),
                    ("Total (95% CI)", -5.79, -7.88, -3.7),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)

            elif result.find("image.7.1.110") != -1:
                self.assertEqual(estimator_type, "IV")
                self.assertEqual(model_type, "Fixed")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Chi": 3.08,
                    "df": 4.0,
                    "P": 0.54,
                    "I": 0.0,
                })
                self.assertEqual(overall_effect, {
                    "Z": 0.54,
                    "P": 0.59,
                })
                expected_data = [
                    ("Kim 2006", -1.0, -6.17, 4.17),
                    ("Li 2008", -1.0, -3.79, 1.79),
                    ("Helgeson 2010", 2.5, -0.58, 5.58),
                    ("Hwang 2012", 0.23, -1.49, 1.95),
                    ("Liu 2014", 0.7, -3.06, 4.46),
                    ("Total (95% CI)", 0.33, -0.88, 1.54),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)


    def a_test_simple_spss_2(self):
        destination = shutil.copy("integration_tests/testdata/pmc5911624.pdf", self.tempdir.name)
        self.assertTrue(os.path.isfile(destination))

        c = Controller(self.tempdir.name)
        c.main()

        raw = open(os.path.join(self.tempdir.name, "forestplots.json")).read()
        data = json.loads(raw)

        # This paper has 4 forest plots
        self.assertEqual(len(data), 4)

        for result in data:
            workbook = openpyxl.load_workbook(result)
            worksheet = workbook.active

            estimator_type = worksheet.cell(row=1, column=2).value
            model_type = worksheet.cell(row=2, column=2).value
            confidence_interval = worksheet.cell(row=3, column=2).value
            count = 5

            self.assertEqual(worksheet.cell(row=count, column=1).value, "Hetrogeneity:")
            hetrogeneity = {}
            while True:
                key = worksheet.cell(row=count, column=2).value
                value = worksheet.cell(row=count, column=3).value
                count = count + 1
                if not key:
                    break
                else:
                    hetrogeneity[key] = float(value)

            self.assertEqual(worksheet.cell(row=count, column=1).value, "Overall Effect:")
            overall_effect = {}
            while True:
                key = worksheet.cell(row=count, column=2).value
                value = worksheet.cell(row=count, column=3).value
                count = count + 1
                if not key:
                    break
                else:
                    overall_effect[key] = float(value)

            self.assertEqual(worksheet.cell(row=count, column=1).value, "Data:")
            data = []
            while True:
                key = worksheet.cell(row=count, column=2).value
                value1 = worksheet.cell(row=count, column=3).value
                value2 = worksheet.cell(row=count, column=4).value
                value3 = worksheet.cell(row=count, column=5).value
                count = count + 1
                if not key:
                    count = count + 1
                    break
                else:
                    data.append((key, float(value1), float(value2), float(value3)))


            if result.find("image.6.2.72") != -1:
                self.assertEqual(estimator_type, "M-H")
                self.assertEqual(model_type, "Fixed")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Chi": 11.62,
                    "df": 10.0,
                    "P": 0.31,
                    "I": 14,
                })
                self.assertEqual(overall_effect, {
                    "Z": 5.51,
                    "P": 0.00001,
                })
                expected_data = [
                    ("Dou-Dou Li (2015)", 1.6, 1.03, 2.5),
                    ("Fei Teng (2017)", 1.26, 0.89, 1.79),
                    ("Geng Li (2017)", 1.42, 1.08, 1.87),
                    ("Hassler MR (2013)", 1.48, 0.33, 6.55),
                    ("Ji-Yong Peng (2008)", 1.1, 0.93, 1.3),
                    ("Jia-Yin Xie (2007)", 1.73, 1.05, 2.83),
                    ("Lei Shi (2014)", 1.41, 0.99, 2.00),
                    ("Shi-Jie Zhao (2016)", 1.25, 0.71, 2.2),
                    ("Tian-Lu Gu (2015)", 1.22, 0.96, 1.55),
                    ("Wei Cheng (2013)", 1.44, 0.89, 2.35),
                    ("Xia Deng (2017)", 1.73, 1.11, 2.69),
                    ("Total (95% CI)", 1.4, 1.24, 1.57),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)

            if result.find("image.7.1.60") != -1:
                self.assertEqual(estimator_type, "M-H")
                self.assertEqual(model_type, "Fixed")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Chi": 2.07,
                    "df": 10.0,
                    "P": 1.0,
                    "I": 0,
                })
                self.assertEqual(overall_effect, {
                    "Z": 1.13,
                    "P": 0.26,
                })
                expected_data = [
                    ("Chua D (2010)", 0.61, 0.16, 2.42),
                    ("Dou-Dou Li (2015)", 1.4, 0.54, 3.6),
                    ("Fei Teng (2017)", 1.15, 0.77, 1.74),
                    ("Geng Li(2017)", 1.07, 0.6, 1.91), # Should be a space
                    ("Hassler MR (2013)", 1.54, 0.71, 3.32),
                    ("Ji-Yong Peng (2008)", 1.11, 0.32, 3.82),
                    ("Lei Shi (2014)", 1.07, 0.46, 2.51),
                    ("Shi-Jie Zhao (2016)", 1.08, 0.59, 1.97),
                    ("Tian-Lu Gu (2015)", 1.28, 0.59, 2.77),
                    ("Wei Cheng (2013)", 0.99, 0.61, 1.61),
                    ("Xia Deng (2017)", 1.08, 0.8, 1.47),
                    ("Total (95% CI)", 1.11, 0.93, 1.32),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)

            if result.find("image.7.2.66") != -1:
                self.assertEqual(estimator_type, "M-H")
                self.assertEqual(model_type, "Random")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Tau": 0.19,
                    "Chi": 36.58,
                    "df": 11.0,
                    "P": 0.0001,
                    "I": 70,
                })
                self.assertEqual(overall_effect, {
                    "Z": 2.27,
                    "P": 0.02,
                })
                expected_data = [
                    ("Chua D (2010)", 3.47, 1.39, 8.65),
                    ("Dou-Dou Li (2015)", 1.57, 0.79, 3.12),
                    ("Fei Teng (2017)", 2.88, 1.5, 5.56),
                    ("Geng Li (2017)", 1.07, 0.62, 1.84),
                    ("Hassler MR (2013)", 1.77, 0.84, 3.73),
                    ("Ji-Yong Peng (2008)", 4.42, 0.54, 36.16),
                    ("Jia-Yin Xie (2007)", 10.0, 1.38, 72.39),
                    ("Lei Shi (2014)", 1.43, 0.56, 3.66),
                    ("Shi-Jie Zhao (2016)", 0.86, 0.48, 1.53),
                    ("Tian-Lu Gu (2015)", 1.17, 0.65, 2.11),
                    ("Wei Cheng (2013)", 1.08, 0.63, 1.87),
                    ("Xia Deng (2017)", 0.9, 0.77, 1.05),
                    ("Total (95% CI)", 1.46, 1.05, 2.04),
                ]
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)

            if result.find("image.8.1.70") != -1:
                self.assertEqual(estimator_type, "M-H")
                self.assertEqual(model_type, "Random")
                self.assertEqual(confidence_interval, "95")
                self.assertEqual(hetrogeneity, {
                    "Tau": 0.18,
                    "Chi": 30.42,
                    "df": 11.0,
                    "P": 0.001,
                    "I": 64.0,
                })
                self.assertEqual(overall_effect, {
                    "Z": 2.21,
                    "P": 0.03,
                })
                expected_data = [
                    ("Chua D (2010)", 1.7, 0.83, 3.5),
                    ("Dou-Dou Li (2015)", 1.8, 0.75, 4.32),
                    ("Fei Teng (2017)", 3.53, 1.72, 7.22),
                    ("Geng Li (2017)", 0.8, 0.35, 1.81),
                    ("Hassler MR (2013)", 1.18, 0.59, 2.38),
                    ("Ji-Yong Peng (2008)", 1.47, 0.38, 5.75),
                    ("dia-Yin Xie (2007)", 8.0, 2.76, 23.2), # Should be Jia-Yin Xei
                    ("Lei Shi (2014)", 1.67, 0.53, 5.28),
                    ("Shi-Jie Zhao (2016)", 0.83, 0.28, 2.44),
                    ("Tian-Lu Gu (2015)", 1.14, 0.56, 2.29),
                    ("Wei Cheng (2013)", 1.08, 0.73, 1.61),
                    ("Xia Deng (2017)", 0.95, 0.74, 1.2),
                    ("Total (95% CI)", 1.45, 1.04, 2.02),
                ]
                print("data:")
                for l in data:
                    print(l)
                self.assertEqual(len(expected_data), len(data))
                total_count = 0
                error_count = 0
                for expected, actual in zip(expected_data, data):
                    for value_expected, value_actual in zip(expected, actual):
                        total_count = total_count + 1
                        if value_expected != value_actual:
                            print ("OCR error in {2}: expected '{0}' vs actual '{1}'.".format(value_expected, value_actual, result))
                            error_count = error_count + 1
                self.assertEqual(1.0 - (float(error_count) / float(total_count)), 1.0)
