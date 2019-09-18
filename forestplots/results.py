"""Module for generating final output."""

import collections

import openpyxl

from forestplots.spssplots import SPSSForestPlot
from forestplots.stataplots import StataForestPlot

ROW_MAJOR_TITLE = 2
ROW_MINOR_TITLE = 3

Header = collections.namedtuple('Header', 'major_header minor_headers')

class Headers:
    def __init__(self, header_list):
        self.header_list = header_list

        column = 0
        for header in header_list:
            for minor_header in header.minor_headers:
                title = 'COLUMN_' + minor_header.upper().replace(' ', '_').replace('-', '_').replace('%', '')
                setattr(self, title, column)
                column += 1

    def width(self):
        return sum([len(x.minor_headers) for x in self.header_list])

OVERALL_HEADERS = Headers([
    Header('', ['Sw type']),
    Header('', ['Forest plot']),
    Header('', ['PDF Image']),
    Header('', ['Has subplots']),
    Header('effect size', ['effect size', 'CI lower bound', 'CI upper bound']),
    Header('weight', ['weight']),
])

SUBGROUP_HEADERS = Headers([
    Header('sub group title', ['title']),
    Header('model type', ['fixed effect', 'random effect']),
    Header('effect size type', ['RR', 'OR', 'SMD', 'WMD']),
    Header('estimator type', ['IV', 'M-H']),
    Header('pooled effect size', ['CI', 'CI lower bound', 'CI upper bound']),
    Header('Confidence interval type', ['90%', '95%', '99%']),
    Header('Heteroegenity statistics', ['Chi-squared', 'df', 'p-value', 'I-squared', 'Tau-squared']),
    Header('test for overall effect', ['Z-value', 'Overall P-value']),
    Header('Group A label', ['a']),
    Header('Group B label', ['b']),
    Header('Direction of effect', ['Favours a', 'Favours b']),
])



class Results:

    def __init__(self, papers):
        self.papers_list = papers

    @staticmethod
    def bordered_cell(worksheet, row, column, value):

        thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                             right=openpyxl.styles.Side(style='thin'),
                                             top=openpyxl.styles.Side(style='thin'),
                                             bottom=openpyxl.styles.Side(style='thin'))

        cell = worksheet.cell(row=row, column=column, value=value)
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        return cell

    @staticmethod
    def plain_cell(worksheet, row, column, value):

        cell = worksheet.cell(row=row, column=column, value=value)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        return cell

    def generate_headers(self, worksheet):

        # need to work out the most subgroups in any paper
        subgroup_max = 1
        for paper in self.papers_list:
            for plot in paper.plots:
                if len(plot.table_list) - 1 > subgroup_max:
                    subgroup_max = len(plot.table_list) - 1

        column = 2
        for header in OVERALL_HEADERS.header_list:
            start_column = column
            self.bordered_cell(worksheet, ROW_MAJOR_TITLE, column, header.major_header)
            for sub_header in header.minor_headers:
                self.bordered_cell(worksheet, ROW_MINOR_TITLE, column, sub_header)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = max(len(sub_header) + 2, (len(header.major_header) + 2) / len(header.minor_headers))
                column += 1

            if start_column < column - 1:
                worksheet.merge_cells(start_row=ROW_MAJOR_TITLE, start_column=start_column,
                                      end_row=ROW_MAJOR_TITLE, end_column=column - 1)


        for _ in range(subgroup_max):
            for header in SUBGROUP_HEADERS.header_list:
                start_column = column
                self.bordered_cell(worksheet, ROW_MAJOR_TITLE, column, header.major_header)
                for sub_header in header.minor_headers:
                    self.bordered_cell(worksheet, ROW_MINOR_TITLE, column, sub_header)
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = max(len(sub_header) + 2, (len(header.major_header) + 2) / len(header.minor_headers))
                    column += 1

                if start_column < column - 1:
                    worksheet.merge_cells(start_row=ROW_MAJOR_TITLE, start_column=start_column,
                                          end_row=ROW_MAJOR_TITLE, end_column=column - 1)



    def save(self, path):
        """Save a workbook containing a summary of all plots."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Summary"


        # Headers
        self.generate_headers(worksheet)

        row = ROW_MINOR_TITLE + 1
        for paper in self.papers_list:

            self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_SW_TYPE, "")
            try:
                self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_FOREST_PLOT, int(paper.pmcid))
            except ValueError:
                self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_FOREST_PLOT, paper.name)
            self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_PDF_IMAGE, "")
            row += 1

            for i, plot in enumerate(paper.plots):
                plot_type = "stata"
                if isinstance(plot, SPSSForestPlot):
                    plot_type = "spss"
                self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_SW_TYPE, plot_type)
                self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_FOREST_PLOT, i)
                self.bordered_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_PDF_IMAGE, plot.id)

                if len(plot.table_list) > 1:
                    self.plain_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_HAS_SUBPLOTS, "x")

                # get the final table
                try:
                    last_table = plot.get_table(-1)
                    last_row = last_table.collapse_data()[-1]
                except IndexError:
                    row += 1
                    continue


                self.plain_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_EFFECT_SIZE, last_row[1])
                self.plain_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_CI_LOWER_BOUND, last_row[2])
                self.plain_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_CI_UPPER_BOUND, last_row[3])

                try:
                    self.plain_cell(worksheet, row, 2 + OVERALL_HEADERS.COLUMN_WEIGHT, last_row[4])
                except IndexError:
                    pass

                subplots = None
                if len(plot.table_list) > 1:
                    subplots = plot.table_list[:-1]
                else:
                    subplots = plot.table_list[:1]
                offset = 2 + OVERALL_HEADERS.width()

                for table in subplots:
                    try:
                        self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_TITLE, table.collapse_titles())
                    except ValueError:
                        pass

                    last_row = table.collapse_data()[-1]
                    self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_CI, last_row[1])
                    self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_CI_LOWER_BOUND, last_row[2])
                    self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_CI_UPPER_BOUND, last_row[3])

                    if isinstance(plot, SPSSForestPlot):
                        for col, key in [
                            (SUBGROUP_HEADERS.COLUMN_CHI_SQUARED, "Chi"),
                            (SUBGROUP_HEADERS.COLUMN_TAU_SQUARED, "Tau"),
                            (SUBGROUP_HEADERS.COLUMN_DF, "df"),
                            (SUBGROUP_HEADERS.COLUMN_P_VALUE, "P"),
                            (SUBGROUP_HEADERS.COLUMN_I_SQUARED, "I")
                        ]:
                            try:
                                self.plain_cell(worksheet, row, offset + col, plot.hetrogeneity[key])
                            except KeyError:
                                pass
                        for col, key in [
                            (SUBGROUP_HEADERS.COLUMN_OVERALL_P_VALUE, "P"),
                            (SUBGROUP_HEADERS.COLUMN_Z_VALUE, "Z"),
                        ]:
                            try:
                                self.plain_cell(worksheet, row, offset + col, plot.overall_effect[key])
                            except KeyError:
                                pass
                        try:
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_A, plot.group_a)
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_B, plot.group_b)
                        except AttributeError:
                            pass

                    elif isinstance(plot, StataForestPlot):

                        for col, key in [
                            (SUBGROUP_HEADERS.COLUMN_I_SQUARED, "i^2"),
                            (SUBGROUP_HEADERS.COLUMN_P_VALUE, "p"),
                        ]:
                            try:
                                self.plain_cell(worksheet, row, offset + col, table.metadata[key])
                            except KeyError:
                                pass

                    try:
                        if plot.mid_point < last_row[1]:
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_FAVOURS_A, "x")
                        if plot.mid_point > last_row[1]:
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_FAVOURS_B, "x")
                    except AttributeError:
                        pass

                    ci_type = plot.summary['Confidence interval']
                    if ci_type == "90":
                        self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_90, "x")
                    elif ci_type == "95":
                        self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_95, "x")
                    elif ci_type == "99":
                        self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_99, "x")

                    # I've mixed estimator type and effective size type
                    try:
                        estimator_type = plot.summary['Esimator type']
                        if estimator_type == 'RR':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_RR, "x")
                        if estimator_type == 'OR':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_OR, "x")
                        if estimator_type == 'SMD':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_SMD, "x")
                        if estimator_type == 'WMD':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_WMD, "x")
                        if estimator_type == 'IV':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_IV, "x")
                        if estimator_type == 'M-H':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_M_H, "x")
                    except KeyError:
                        pass

                    try:
                        model_type = plot.summary['Model type']
                        if model_type == 'Fixed':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_FIXED_EFFECT, "x")
                        if model_type == 'Random':
                            self.plain_cell(worksheet, row, offset + SUBGROUP_HEADERS.COLUMN_RANDOM_EFFECT, "x")
                    except KeyError:
                        pass


                    offset += SUBGROUP_HEADERS.width()


                row += 1


        workbook.save(path)
