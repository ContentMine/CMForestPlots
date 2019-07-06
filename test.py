#!/usr/bin/env python3

import os
import re
import subprocess
import sys
import xml.dom.minidom

import openpyxl

USE_DOCKER = True
IMAGE_NAME = "forestplot"

# Heterogeneity: Tau? = 0.00; Chi? = 2.98, df= 4 (P = 0.56), 7= 0% |\nTest for overall effect: Z= 3.12
# b'Heterogeneity: Chi? = 2.07, df= 10 (P= 1.00); /7= 0%\nTest for overall effect: Z= 1.13 (P = 0.26)\n\x0c'
# b'Heterogeneity: Ch? = 2.11, df = 5 (P = 0.83); P= 0%\nTest for overall effect: Z = 3.80 (P = 0.0001)\n\n \n\x0c'

def forgiving_float(s):
    return float(s.replace('~', '-').replace(',', '.'))

class Paper(object):

    def __init__(self, ctree_directory):
        self.ctree_directory = ctree_directory
        self.plots = []

class ForestPlot(object):

    def __init__(self, image_directory):
        self.image_directory = image_directory
        self.tau = None
        self.chi = None
        self.df = None
        self.P  = None
        self.I = None
        self.Z = None
        self.ZP = None
        self.OverallZ = None
        self.OverallP = None
        self.titles = []
        self.values = []

    def add_footer_summery(self, footer):
        pass

    def is_valid(self):
        return len(self.titles) > 0

    def save(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"


        ws.cell(row=1, column=1, value="Hetrogeneity:")
        count = 1
        if self.tau:
            ws.cell(row=count, column=2, value="Tau")
            ws.cell(row=count, column=3, value=self.tau)
            count += 1
        if self.chi:
            ws.cell(row=count, column=2, value="Chi")
            ws.cell(row=count, column=3, value=self.chi)
            count += 1
        if self.df:
            ws.cell(row=count, column=2, value="DF")
            ws.cell(row=count, column=3, value=self.df)
            count += 1
        if self.P:
            ws.cell(row=count, column=2, value="P")
            ws.cell(row=count, column=3, value=self.P)
            count += 1
        if self.I:
            ws.cell(row=count, column=2, value="I2")
            ws.cell(row=count, column=3, value=self.I)
            count += 1
        count += 1

        ws.cell(row=count, column=1, value="Overall Effect:")
        if self.OverallZ:
            ws.cell(row=count, column=2, value="Z")
            ws.cell(row=count, column=3, value=self.OverallZ)
            count += 1
        if self.OverallP:
            ws.cell(row=count, column=2, value="P")
            ws.cell(row=count, column=3, value=self.OverallP)
            count += 1
        count += 1

        ws.cell(row=count, column=1, value="Data:")
        for title, value in zip(self.titles, self.values):
            ws.cell(row=count, column=2, value=title)
            ws.cell(row=count, column=3, value=value[0])
            ws.cell(row=count, column=4, value=value[1])
            ws.cell(row=count, column=5, value=value[2])
            count += 1

        wb.save(os.path.join(self.image_directory, "results.xlsx"))


class SPSSForestPlot(ForestPlot):
    SUMMARY_RE = re.compile("^\s*Heterogeneity:(\s*Tau.\s*=\s*|)([\d.,]*)(\s*[,;]|)\s*Chi.\s*=\s*([\d.,]+)\s*[,;]\s*df\s*=\s*([\d.,]+)\s*\(\s*P\s*=\s*([\d,.]+)\s*\)\s*[,;]\s*[IPF7]\s*=\s*([\d.,]+%)[\s\n]*Test for overall effect:\s*Z\s*=\s*([\d.,]+)\s*\(\s*P\s*=\s*([\d.,]+)\s*\)")


class StataForestPlot(ForestPlot):
    pass



class Controller(object):

    def __init__(self, project_directory):
        self.project_directory = project_directory

    def docker_wrapper(self, command, args):
        subprocess.run(["docker", "run", "-it", "--rm", "-v", "{0}:/tmp/project".format(self.project_directory),
            IMAGE_NAME, command, "-p", "/tmp/project"] + args, capture_output=True)

    def normami_command(self, command, args=None):
        if not args:
            args = []
        print("Calling {0}".format(command))
        if USE_DOCKER:
            self.docker_wrapper(command, args)
        else:
            subprocess.run([command, "-p", self.project_directory] + args, capture_output=True)
        print("Done")


    def main(self):

        if not os.path.isfile(os.path.join(self.project_directory, "make_project.json")):
            print("Generating CProject in {0}...".format(self.project_directory))
            self.normami_command("ami-makeproject", ["--rawfiletypes", "html,pdf,xml", "--omit", "template.xml"])

        raw_project_contents = [os.path.join(self.project_directory, x) for x in os.listdir(self.project_directory)]
        project_contents = [x for x in raw_project_contents if os.path.isdir(x)]

        needs_pdf_parsing = False
        for ctree in project_contents:
            if not os.path.isdir(os.path.join(ctree, "pdfimages")):
                needs_pdf_parsing = True
                break
        if needs_pdf_parsing:
            self.normami_command("ami-pdf")
            for threshold in ["150"]:
                print("Testing with threshold {0}".format(threshold))
                self.normami_command("ami-image", ["--sharpen", "sharpen4", "--threshold", threshold, "--despeckle", "true"])
                self.normami_command("ami-pixel", ["--projections", "--yprojection", "0.4",
                    "--xprojection", "0.7", "--lines", "--minheight", "1", "--rings", "-1", "--islands", "0",
                    "--inputname", "raw_s4_thr_{0}_ds".format(threshold),
                    "--templateinput", "raw_s4_thr_{0}_ds/projections.xml".format(threshold),
                    "--templateout", "template.xml",
                    "--templatexsl", "/org/contentmine/ami/tools/spssTemplate.xsl"])

                self.normami_command("ami-forestplot", ["--segment", "--template", "raw_s4_thr_{0}_ds/template.xml".format(threshold)])

        papers = []
        for ctree in project_contents:
            paper = Paper(ctree)
            papers.append(paper)

            pdf_images_dir = os.path.join(ctree, "pdfimages")
            imagedirs = [os.path.join(pdf_images_dir, x) for x in os.listdir(pdf_images_dir) if x.startswith("image.")]
            for imagedir in imagedirs:
                print(imagedir)

                plot = ForestPlot(imagedir)

                image_path = os.path.join(imagedir, "raw.footer.summary.png")
                if not os.path.isfile(image_path):
                    continue

                best_res = None
                for threshold in range(50, 80, 2):
                    output_ocr_name = os.path.join(imagedir, "footer.summary.{0}.txt".format(threshold))
                    if not os.path.isfile(output_ocr_name):
                        output_image_name = os.path.join(imagedir, "footer.summary.{0}.png".format(threshold))
                        if not os.path.isfile(output_image_name):
                            subprocess.run(["convert", "-black-threshold", "{0}%".format(threshold), image_path, output_image_name])
                        subprocess.run(["tesseract", output_image_name, os.path.splitext(output_ocr_name)[0]], capture_output=True)

                    ocr_prose = open(output_ocr_name).read()
                    m = SPSSForestPlot.SUMMARY_RE.match(ocr_prose)
                    if m:
                        if not best_res:
                            best_res = m.groups()
                        else:
                            if sum([1 for x in m.groups() if len(x) > 0]) > sum([1 for x in best_res if len(x) > 0]):
                                best_res = m.groups()

                if best_res:
                    try:
                        plot.tau = forgiving_float(best_res[1])
                    except ValueError:
                        pass
                    try:
                        plot.chi = forgiving_float(best_res[3])
                    except ValueError:
                        pass
                    try:
                        plot.df = forgiving_float(best_res[4])
                    except ValueError:
                        pass
                    try:
                        plot.P = forgiving_float(best_res[5])
                    except ValueError:
                        pass
                    try:
                        plot.I = forgiving_float(best_res[6])
                    except ValueError:
                        pass
                    try:
                        plot.OverallZ = forgiving_float(best_res[7])
                    except ValueError:
                        pass
                    try:
                        plot.OverallP = forgiving_float(best_res[8])
                    except ValueError:
                        pass

                image_path = os.path.join(imagedir, "raw.body.table.png")
                if not os.path.isfile(image_path):
                    continue

                best_res = None
                for threshold in range(50, 80, 2):
                    print("Threshold {0}".format(threshold))
                    output_ocr_name = os.path.join(imagedir, "body.table.{0}.txt".format(threshold))
                    if not os.path.isfile(output_ocr_name):
                        output_image_name = os.path.join(imagedir, "body.table.{0}.png".format(threshold))
                        if not os.path.isfile(output_image_name):
                            subprocess.run(["convert", "-black-threshold", "{0}%".format(threshold), image_path, output_image_name])
                        subprocess.run(["tesseract", output_image_name, os.path.splitext(output_ocr_name)[0]], capture_output=True)

                    titles = []
                    raw_lines = open(output_ocr_name, 'r').readlines()
                    lines = [x for x in raw_lines if len(x.strip()) > 0]
                    print("We got {0} lines".format(len(lines)))

                    for line in lines:
                        if line.startswith('Total'):
                            titles.append(line)
                            break
                        titles.append(line)
                    print("We got {0} titles.".format(len(titles)))

                    values = []
                    r = re.compile('^\s*([-~]{0,1}\d+[.,]{0,1}\d*)\s*[\[\({]([-~]{0,1}\d+[.,]{0,1}\d*)\s*,\s*([-~]{0,1}\d+[.,]{0,1}\d*)[\]}\)]\s*$')
                    for line in lines:
                        m = r.match(line)
                        if m:
                            g = m.groups()
                            try:
                                values.append((forgiving_float(g[0]), forgiving_float(g[1]), forgiving_float(g[2])))
                            except ValueError:
                                pass
                    print("We got {0} values.".format(len(values)))

                    if len(values) > 0 and len(values) == len(titles):
                        if not best_res:
                            best_res = (titles, values)
                        else:
                            if len(titles) > len(best_res[0]):
                                best_res = (titles, values)
                if best_res:
                    plot.titles = best_res[0]
                    plot.values = best_res[1]

                if plot.is_valid():
                    paper.plots.append(plot)
                plot.save()

        for paper in papers:
            print("Paper {1} has {0} plots.".format(len(paper.plots), paper.ctree_directory))
            for plot in paper.plots:
                print("\t{0}".format(plot.image_directory))




if __name__ == "__main__":

    if len(sys.argv) != 2:
        print("USAGE: {0} PROJECT_DIRECTORY".format(sys.argv[0]))
        sys.exit(-1)

    if not os.path.isdir(sys.argv[1]):
        print("USAGE: {0} PROJECT_DIRECTORY".format(sys.argv[0]))
        sys.exit(-1)

    c = Controller(sys.argv[1])
    c.main()
